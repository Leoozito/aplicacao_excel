import openpyxl
import os
import pyautogui
import subprocess
import pyperclip
import time

# Pegar dados do Excel
arquivo_excel = openpyxl.load_workbook("arquivo_excel.xlsx")
dados = arquivo_excel.active

total_linhas = dados.max_row
total_colunas = dados.max_column

# Caminho onde será salvo o arquivo
desktop_path = os.path.expanduser("~/Área de Trabalho")
nome_arquivo = "relatorio.txt"
file_path = os.path.join(desktop_path, nome_arquivo)

lista_valores = []  # Lista que será randomizada na calculadora

with open(file_path, "w") as file:
    for linha in range(1, total_linhas + 1):  # para manipular cada linha
        for coluna in range(1, total_colunas + 1):  # para manipular cada coluna
            celula = dados.cell(row=linha, column=coluna)
            valor = celula.value

            if valor is not None:
                conteudo = f"linha {linha}, coluna {coluna}: {valor}\n"
                if type(valor) == int:
                    lista_valores.append(valor)
                file.write(conteudo)  # adiciona as informações do Excel no arquivo

# Abrir os softwares
os.system(f"xdg-open '{file_path}'")

subprocess.Popen(["gnome-calculator"])
for valor in lista_valores:
    pyautogui.press('+')
    time.sleep(0.5)  # Aguarde meio segundo entre o pressionamento de '+' e a entrada do valor
    pyautogui.write(str(valor))
    pyautogui.press('enter')

pyautogui.hotkey('ctrl', 'c')
resultado = pyperclip.paste()
with open(file_path, 'a') as file:  # Use 'a' para modo de anexação
    file.write(resultado)

célula_resultado = dados.cell(row=23, column=2)  # Por exemplo, adicione à primeira linha da última coluna
célula_resultado.value = resultado

# Salve as alterações no arquivo Excel
arquivo_excel.save("arquivo_excel.xlsx")
pyautogui.hotkey('f5')
os.system(f"xdg-open 'arquivo_excel.xlsx'")
