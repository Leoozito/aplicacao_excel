import openpyxl
import os
import pyautogui
import subprocess


# Pegar dados do Excel
arquivo_excel = openpyxl.load_workbook("arquivo_excel.xlsx")
dados = arquivo_excel.active

total_linhas = dados.max_row
total_colunas = dados.max_column

# Caminho onde será salvo o arquivo
desktop_path = os.path.expanduser("~/Área de Trabalho")
nome_arquivo = "relatorio.txt"
file_path = os.path.join(desktop_path, nome_arquivo)

lista_valores = [] # Lista que será randomizada na calculadora

with open(file_path, "w") as file:
    for linha in range(1, total_linhas + 1): # para manipular cada linha
        for coluna in range(1, total_colunas + 1): # para manipular cada coluna
            celula = dados.cell(row=linha, column=coluna)
            valor = celula.value
            if valor is not None:
                conteudo = f"linha {linha}, coluna {coluna}: {valor}\n"
                if type(valor) == int:
                    lista_valores.append(valor)                    
                file.write(conteudo) # adiciona as informações do Excel no arquivo

# Abrir os softwares
os.system(f"xdg-open '{file_path}'")

subprocess.Popen(["gnome-calculator"])

for valor in lista_valores: 
    pyautogui.write(str(valor))  
    pyautogui.press('enter')        

