import openpyxl
import os
import pyautogui
import subprocess
import threading
import time  # Importe o módulo time

# Pegar dados do Excel
arquivo_excel = openpyxl.load_workbook("arquivo_excel.xlsx")
dados = arquivo_excel.active

total_linhas = dados.max_row
total_colunas = dados.max_column

# Caminho onde será salvo o arquivo (Área de Trabalho)
desktop_path = os.path.expanduser("~/Documents")
nome_arquivo = "relatorio.txt"
file_path = os.path.join(desktop_path, nome_arquivo)

lista_valores = []  # Lista que será randomizada na calculadora

def open_notepad():
    os.system(f"notepad.exe {file_path}")

def open_calculator():
    calculator_process = subprocess.Popen(["calc.exe"])
    calculator_title = "Calculadora"

    time.sleep(5)  # Aguarde um tempo maior (5 segundos) para a janela da calculadora abrir

    calculator_windows = pyautogui.getWindowsWithTitle(calculator_title)

    if calculator_windows:  # Verifique se a lista não está vazia
        calculator_window = calculator_windows[0]

        calculator_window.activate()

        for valor in lista_valores:
            pyautogui.write(str(valor))
            pyautogui.press('enter')

        calculator_process.wait()

with open(file_path, "w") as file:
    for linha in range(1, total_linhas + 1):  # para manipular cada linha
        for coluna in range(1, total_colunas + 1):  # para manipular cada coluna
            celula = dados.cell(row=linha, column=coluna)
            valor = celula.value
            if valor is not None:
                conteudo = f"linha {linha}, coluna {coluna}: {valor}\n"
                if isinstance(valor, int):
                    lista_valores.append(valor)
                file.write(conteudo)  # adiciona as informações do Excel no arquivo

# Criação das threads
notepad_thread = threading.Thread(target=open_notepad)
calculator_thread = threading.Thread(target=open_calculator)

# Inicia as threads
notepad_thread.start()
calculator_thread.start()

# Aguarda as threads terminarem
notepad_thread.join()
calculator_thread.join()